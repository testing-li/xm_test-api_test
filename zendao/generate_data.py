# -*- coding: utf-8 -*-
from jinja2 import Environment, FileSystemLoader
from openpyxl import Workbook


def generate_bug_html(dever_data, tester_data, date, name):
    """
    通过template.html模板生成html文件
    :param data: 禅道源页面json格式数据
    :return:
    """
    env = Environment(loader=FileSystemLoader('./'))
    template = env.get_template('template.html')
    with open("./result.html", 'w+', encoding='utf-8') as f:
        html_content = template.render(time=date,
                                       name=name,
                                       num=dever_data['total_num'],
                                       solve_num=dever_data['total_solve_num'],
                                       active_num=dever_data['total_active_num'],
                                       dever_data=dever_data["rows"],
                                       tester_data=tester_data)
        f.write(html_content)


def generate_bug_excel(data, write_path):
    """
    通过bug列表数据生成excel表格文件
    :param data: 禅道源页面json格式数据
    :param write_path: 生成文件路径
    :return:
    """
    title = ["Bug编号", "严重级别", "优先级", "Bug标题", "Bug状态", "由谁创建", "指派给",
             "解决者", "解决方案", "归属者", "归属说明", "代码审核人"]
    severity = {
        "1": "致命错误",
        "2": "严重错误",
        "3": "一般错误",
        "4": "较小错误",
        "5": "建议错误"
    }
    pri = {
        "1": "立即",
        "2": "急需",
        "3": "高",
        "4": "中",
        "5": "低",
    }
    status = {
        "active": "未解决",
        "closed": "已关闭",
        "resolved": "已解决"
    }
    resolution = {
        'bydesign': "设计如此",
        'duplicate': "重复Bug",
        'external': "外部原因",
        'fixed': "已解决",
        'notrepro': "无法重现",
        'postponed': "延期处理",
        'willnotfix': "不予解决",
        'tostory': "转为需求"
    }
    wb = Workbook()
    ws = wb.active
    ws.title = "线上bug"
    ws1 = wb.create_sheet("测试环境bug")
    test_bug_list = data["other_bugs"]["bugs"]
    master_bug_list = data["master_bugs"]["bugs"]
    for i in range(1, len(title) + 1):
        ws.cell(row=1, column=i, value=title[i - 1])
        ws1.cell(row=1, column=i, value=title[i - 1])
    for i in range(len(master_bug_list)):
        ws.cell(row=i + 2, column=1, value=master_bug_list[i]['id'])
        ws.cell(row=i + 2, column=2, value=severity.get(master_bug_list[i]['severity'], master_bug_list[i]['severity']))
        ws.cell(row=i + 2, column=3, value=pri.get(master_bug_list[i]['pri'], master_bug_list[i]['pri']))
        ws.cell(row=i + 2, column=4, value=master_bug_list[i]['title'])
        ws.cell(row=i + 2, column=5, value=status.get(master_bug_list[i]['status'], master_bug_list[i]['status']))
        ws.cell(row=i + 2, column=6,
                value=data['users'].get(master_bug_list[i]['openedBy'], master_bug_list[i]['openedBy']))
        ws.cell(row=i + 2, column=7, value=data['users'][master_bug_list[i]['assignedTo']])
        ws.cell(row=i + 2, column=8, value=data['users'].get(master_bug_list[i]['resolvedBy']))
        ws.cell(row=i + 2, column=9,
                value=resolution.get(master_bug_list[i]['resolution'], master_bug_list[i]['resolution']))
    for i in range(len(test_bug_list)):
        ws1.cell(row=i + 2, column=1, value=test_bug_list[i]['id'])
        ws1.cell(row=i + 2, column=2, value=severity.get(test_bug_list[i]['severity'], test_bug_list[i]['severity']))
        ws1.cell(row=i + 2, column=3, value=pri.get(test_bug_list[i]['pri'], test_bug_list[i]['pri']))
        ws1.cell(row=i + 2, column=4, value=test_bug_list[i]['title'])
        ws1.cell(row=i + 2, column=5, value=status.get(test_bug_list[i]['status'], test_bug_list[i]['status']))
        ws1.cell(row=i + 2, column=6,
                value=data['users'].get(test_bug_list[i]['openedBy'], test_bug_list[i]['openedBy']))
        ws1.cell(row=i + 2, column=7, value=data['users'].get(test_bug_list[i]['assignedTo']))
        ws1.cell(row=i + 2, column=8, value=data['users'].get(test_bug_list[i]['resolvedBy']))
        ws1.cell(row=i + 2, column=9,
                value=resolution.get(test_bug_list[i]['resolution'], test_bug_list[i]['resolution']))
    wb.save(write_path)
