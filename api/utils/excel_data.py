#!/usr/bin/env python
# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl import load_workbook
import random
import time

'''
# 创建表单
wb = Workbook()  # 实例化
ws1 = wb.active  # 默认创建获取第一个工作表的索引对象
ws2 = wb.create_sheet('mysheet')  # 默认插入到最后，新建一个sheet对象
ws3 = wb.create_sheet('onesheet',0)  # 通过索引改变插入位置
wb.create_sheet()   # 新建一个sheet 名称不指定的话默认Sheet Sheet1 ...创建
ws4 = wb['Sheet1']  # 通过名称所谓索引指定sheet的引用
ws4.title = 'test'  # 重命名
print(wb.sheetnames)
------------------------------------------------------------------------
# 读取文件                  
wb = load_workbook(path)
ws = wb[sheetname]      
------------------------------------------------------------------------
# 访问单元格
d = ws.cell(row=4, column=2, value=10)
# 创建访问单元格并赋值
for x in range(1,101):  
     for y in range(1,101):
 		ws.cell(row=x,colum=y)
# 批量访问单元格对象
# 所有行 ws.rows
# 所有列 ws.colums
# 赋值 c.value = value
-------------------------------------------------------------------------
# 保存文件
# 普通文件 wb.save(name)
# 流文件 wb.save(name,as_template=True)
-------------------------------------------------------------------------
'''


def _phone(n):
    return f'{15400000000 + n}'


def _time(format="%Y/%m/%d"):
    return time.strftime(format, time.localtime(int(round(time.time())) * random.random()))


def _email(n):
    return f'test{n}@tsidy.com'


class Excel:
    '''
    集成method中的数据生成 method.phone()
    str：f"text{n}" or "text"
    int: random.randint(start,end)
    time: random_time
    date: _time(format="%Y/%m/%d %H:%M:%S")
    phone: _phone(n)
    email: _email(n)
    '''

    def read_excel(self, path, sheetname='Sheet1'):
        wb = load_workbook(path)
        ws = wb[sheetname]
        l = []
        for row in ws.values:
            l.append(row)
        return l

    def write_excel(self, read_path, write_path, num, sheetname='Sheet1'):
        wb = Workbook()
        ws = wb.active
        ws.name = sheetname
        demo = self.read_excel(read_path)
        title = demo[0]
        data_type = demo[1]
        for i in range(1, len(title) + 1):
            ws.cell(row=1, column=i, value=title[i - 1])
        for i in range(len(data_type)):
            if not data_type[i]:
                continue
            else:
                for n in range(num):
                    ws.cell(row=n + 2, column=i + 1, value=eval(str(data_type[i])))
        wb.save(write_path)


if __name__ == '__main__':
    Excel().write_excel(r'C:\Users\87859\Desktop\短信投放名单模板.xlsx', r"C:\Users\87859\Desktop\demo_data.xlsx", 20001)
